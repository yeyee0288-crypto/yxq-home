import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os


class ExcelSyncTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 数据互通工具 (精确匹配版)")
        self.root.geometry("800x600")
        
        self.source_file = None
        self.target_file = None
        self.source_wb = None
        self.target_wb = None
        self.source_ws = None
        self.target_ws = None
        self.matching_columns = {}
        
        self.create_ui()
    
    def create_ui(self):
        """创建用户界面"""
        # 文件选择区域
        frame_files = ttk.LabelFrame(self.root, text="文件选择", padding=10)
        frame_files.pack(fill="x", padx=10, pady=10)
        
        # 源文件
        ttk.Label(frame_files, text="源文件 (A):").grid(row=0, column=0, sticky="w")
        self.source_entry = ttk.Entry(frame_files, width=50)
        self.source_entry.grid(row=0, column=1, padx=5)
        ttk.Button(frame_files, text="浏览", command=self.select_source).grid(row=0, column=2)
        
        # 目标文件
        ttk.Label(frame_files, text="目标文件 (B):").grid(row=1, column=0, sticky="w", pady=10)
        self.target_entry = ttk.Entry(frame_files, width=50)
        self.target_entry.grid(row=1, column=1, padx=5)
        ttk.Button(frame_files, text="浏览", command=self.select_target).grid(row=1, column=2)
        
        # 加载按钮
        ttk.Button(frame_files, text="加载文件", command=self.load_files).grid(row=2, column=0, columnspan=3, pady=10)
        
        # 列匹配区域
        frame_columns = ttk.LabelFrame(self.root, text="列匹配信息 (仅显示完全匹配的列)", padding=10)
        frame_columns.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 创建表格
        columns = ("源列 (A)", "→", "目标列 (B)")
        self.tree = ttk.Treeview(frame_columns, columns=columns, height=15, show="headings")
        
        self.tree.column("源列 (A)", width=250, anchor="center")
        self.tree.column("→", width=50, anchor="center")
        self.tree.column("目标列 (B)", width=250, anchor="center")
        
        self.tree.heading("源列 (A)", text="源列 (A)")
        self.tree.heading("→", text="→")
        self.tree.heading("目标列 (B)", text="目标列 (B)")
        
        self.tree.pack(fill="both", expand=True)
        
        # 操作按钮区域
        frame_buttons = ttk.Frame(self.root)
        frame_buttons.pack(fill="x", padx=10, pady=10)
        
        ttk.Button(frame_buttons, text="自动匹配 (完全相同)", command=self.auto_match_exact).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="查看映射", command=self.view_mapping).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="执行同步", command=self.sync_data).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="保存文件", command=self.save_file).pack(side="left", padx=5)
        ttk.Button(frame_buttons, text="清空", command=self.clear_all).pack(side="left", padx=5)
        
        # 状态栏
        self.status_label = ttk.Label(self.root, text="就绪", relief="sunken")
        self.status_label.pack(fill="x", padx=10, pady=5)
    
    def select_source(self):
        """选择源文件"""
        file = filedialog.askopenfilename(
            title="选择源文件 (A)",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file:
            self.source_file = file
            self.source_entry.delete(0, tk.END)
            self.source_entry.insert(0, file)
            self.status_label.config(text=f"已选择源文件: {os.path.basename(file)}")
    
    def select_target(self):
        """选择目标文件"""
        file = filedialog.askopenfilename(
            title="选择目标文件 (B)",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file:
            self.target_file = file
            self.target_entry.delete(0, tk.END)
            self.target_entry.insert(0, file)
            self.status_label.config(text=f"已选择目标文件: {os.path.basename(file)}")
    
    def load_files(self):
        """加载文件"""
        if not self.source_file or not self.target_file:
            messagebox.showerror("错误", "请先选择源文件和目标文件")
            return
        
        try:
            # 加载源文件
            self.source_wb = openpyxl.load_workbook(self.source_file)
            self.source_ws = self.source_wb.active
            
            # 加载目标文件
            self.target_wb = openpyxl.load_workbook(self.target_file)
            self.target_ws = self.target_wb.active
            
            self.status_label.config(text="文件已加载成功")
            messagebox.showinfo("成功", "文件加载成功！")
        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")
            self.status_label.config(text="加载失败")
    
    def get_column_names(self, ws):
        """获取工作表的列名（第一行）"""
        columns = []
        for cell in ws[1]:
            if cell.value is not None:
                columns.append(str(cell.value).strip())
        return columns
    
    def auto_match_exact(self):
        """自动匹配 - 只匹配名称完全相同的列"""
        if not self.source_ws or not self.target_ws:
            messagebox.showerror("错误", "请先加载文件")
            return
        
        # 获取列名
        source_columns = self.get_column_names(self.source_ws)
        target_columns = self.get_column_names(self.target_ws)
        
        # 清空之前的数据
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.matching_columns = {}
        
        # 精确匹配 - 只匹配完全相同的列名
        for source_col in source_columns:
            if source_col in target_columns:
                self.matching_columns[source_col] = source_col
                self.tree.insert("", "end", values=(source_col, "→", source_col))
        
        if self.matching_columns:
            self.status_label.config(text=f"找到 {len(self.matching_columns)} 个完全匹配的列")
            messagebox.showinfo("匹配完成", f"找到 {len(self.matching_columns)} 个完全匹配的列")
        else:
            self.status_label.config(text="未找到完全匹配的列")
            messagebox.showwarning("提示", "未找到完全匹配的列名，请检查列名是否一致")
    
    def view_mapping(self):
        """查看映射关系"""
        if not self.matching_columns:
            messagebox.showinfo("提示", "还没有进行匹配")
            return
        
        mapping_text = "列映射关系:\n\n"
        for source, target in self.matching_columns.items():
            mapping_text += f"{source} → {target}\n"
        
        messagebox.showinfo("映射关系", mapping_text)
    
    def copy_cell_format(self, source_cell, target_cell):
        """复制单元格格式"""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    
    def sync_data(self):
        """执行数据同步"""
        if not self.source_ws or not self.target_ws:
            messagebox.showerror("错误", "请先加载文件")
            return
        
        if not self.matching_columns:
            messagebox.showerror("错误", "请先进行列匹配")
            return
        
        try:
            source_columns = self.get_column_names(self.source_ws)
            target_columns = self.get_column_names(self.target_ws)
            
            # 获取源列和目标列的索引
            source_col_indices = {col: idx + 1 for idx, col in enumerate(source_columns)}
            target_col_indices = {col: idx + 1 for idx, col in enumerate(target_columns)}
            
            # 同步数据
            sync_count = 0
            
            for source_col_name, target_col_name in self.matching_columns.items():
                if source_col_name not in source_col_indices or target_col_name not in target_col_indices:
                    continue
                
                source_col_idx = source_col_indices[source_col_name]
                target_col_idx = target_col_indices[target_col_name]
                
                # 获取源列的最大行数
                max_rows = self.source_ws.max_row
                
                # 复制数据（从第2行开始，跳过标题行）
                for row in range(2, max_rows + 1):
                    source_cell = self.source_ws.cell(row=row, column=source_col_idx)
                    target_cell = self.target_ws.cell(row=row, column=target_col_idx)
                    
                    # 复制值
                    if source_cell.value is not None:
                        target_cell.value = source_cell.value
                        
                        # 复制格式
                        self.copy_cell_format(source_cell, target_cell)
                        
                        sync_count += 1
            
            self.status_label.config(text=f"同步完成，共复制 {sync_count} 个单元格")
            messagebox.showinfo("成功", f"数据同步完成！\n共复制 {sync_count} 个单元格的数据")
        
        except Exception as e:
            messagebox.showerror("错误", f"同步失败: {str(e)}")
            self.status_label.config(text="同步失败")
    
    def save_file(self):
        """保存文件"""
        if not self.target_wb:
            messagebox.showerror("错误", "请先加载目标文件")
            return
        
        try:
            file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            if file:
                self.target_wb.save(file)
                self.status_label.config(text=f"文件已保存: {os.path.basename(file)}")
                messagebox.showinfo("成功", f"文件已保存:\n{file}")
        except Exception as e:
            messagebox.showerror("错误", f"保存文件失败: {str(e)}")
            self.status_label.config(text="保存失败")
    
    def clear_all(self):
        """清空所有数据"""
        self.source_file = None
        self.target_file = None
        self.source_wb = None
        self.target_wb = None
        self.source_ws = None
        self.target_ws = None
        self.matching_columns = {}
        
        self.source_entry.delete(0, tk.END)
        self.target_entry.delete(0, tk.END)
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.status_label.config(text="已清空所有数据")
        messagebox.showinfo("清空完成", "所有数据已清空")


def main():
    root = tk.Tk()
    app = ExcelSyncTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
